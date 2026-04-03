from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .config import DATA_DIR, KeywordQuerySpec

# Excel column width (approx. character units). Unknown headers get _DEFAULT_COL_WIDTH.
_DEFAULT_COL_WIDTH = 20.0
_COLUMN_WIDTHS: dict[str, float] = {
    "Article No.": 11.0,
    "PMID": 11.0,
    "PubMed URL": 12.0,
    "Title": 52.0,
    "Publication Date": 16.0,
    "Journal": 36.0,
    "ISSN": 14.0,
    "Corresponding Author": 28.0,
    "Corresponding Author Affiliation": 48.0,
    "Institution Country": 20.0,
    "Keywords": 40.0,
    "Abstract": 88.0,
    "Abstract (Chinese)": 52.0,
}

# Wrap + top-align for cells that often hold long text (short ID columns stay single-line)
_WRAP_COLUMNS: frozenset[str] = frozenset(
    {
        "Title",
        "Journal",
        "Corresponding Author",
        "Corresponding Author Affiliation",
        "Institution Country",
        "Keywords",
        "Abstract",
        "Abstract (Chinese)",
    }
)


def _apply_excel_readability(ws, *, header_row: int) -> None:
    """Set column widths for all headers; wrap long-text columns."""
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is not None and str(v).strip():
            headers[str(v).strip()] = col

    wrap_align = Alignment(wrap_text=True, vertical="top")

    for name, col_idx in headers.items():
        letter = get_column_letter(col_idx)
        width = _COLUMN_WIDTHS.get(name, _DEFAULT_COL_WIDTH)
        ws.column_dimensions[letter].width = width

        use_wrap = name in _WRAP_COLUMNS or name not in _COLUMN_WIDTHS
        if not use_wrap:
            continue
        for row in range(header_row, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).alignment = wrap_align

    if ws.max_row > header_row:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


_LINK_FONT = Font(color="0563C1", underline="single")


def _apply_pubmed_hyperlinks(ws, *, header_row: int) -> None:
    """Make PubMed URL column clickable (opens in browser from Excel)."""
    col_idx: int | None = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip() == "PubMed URL":
            col_idx = c
            break
    if col_idx is None:
        return
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        raw = cell.value
        if not raw or not isinstance(raw, str):
            continue
        url = raw.strip()
        if not (url.startswith("http://") or url.startswith("https://")):
            continue
        cell.hyperlink = url
        cell.value = "PubMed"
        cell.font = _LINK_FONT


def build_run_summary_lines(
    *,
    mindate: str,
    maxdate: str,
    days: int,
    term: str,
    qspec: KeywordQuerySpec,
    kw_num: int | None,
    use_keywords_json: bool,
    kw_md_basename: str | None = None,
) -> tuple[str, str]:
    """First two rows in Excel: PDAT range + keyword source and PubMed term (English)."""
    line1 = (
        f"PDAT search range: {mindate} — {maxdate} ({days} calendar days, inclusive)"
    )
    if use_keywords_json:
        src = "Keyword source: data/keywords.json (pubmed_query or keywords array)"
    elif kw_md_basename:
        src = f"Keyword file: {kw_md_basename} (data/keywords/{kw_md_basename})"
    else:
        src = f"Keyword file: kw_{kw_num}.md (data/keywords/kw_{kw_num}.md)"
    if qspec.freeform:
        body = (qspec.parts[0] or "").strip() or "(empty)"
        kw_desc = f"Freeform query: {body}"
    else:
        parts = [p.strip() for p in qspec.parts if p.strip()]
        body = "; ".join(parts) if parts else "(empty)"
        kw_desc = f"Keyword lines (md): {body}  |  Between lines: {qspec.join_between}"
    line2 = f"{src}  |  {kw_desc}  |  PubMed full term: {term}"
    return line1, line2


def export_rows(
    rows: list[dict],
    output_path: Path | None = None,
    *,
    run_summary: tuple[str, str] | None = None,
) -> Path:
    if output_path is None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = DATA_DIR / "runs" / f"export_{ts}"
        out_dir.mkdir(parents=True, exist_ok=True)
        output_path = out_dir / "pubmed.xlsx"
    else:
        output_path = Path(output_path).expanduser().resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    preferred = [
        "Article No.",
        "PMID",
        "PubMed URL",
        "Title",
        "Publication Date",
        "Journal",
        "ISSN",
        "Corresponding Author",
        "Corresponding Author Affiliation",
        "Institution Country",
        "Keywords",
        "Abstract",
        "Abstract (Chinese)",
    ]
    cols = [c for c in preferred if c in df.columns]
    rest = [c for c in df.columns if c not in cols]
    df = df[cols + rest]

    sheet_name = "Articles"

    if run_summary is None:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = load_workbook(output_path)
        ws0 = wb[sheet_name]
        _apply_excel_readability(ws0, header_row=1)
        _apply_pubmed_hyperlinks(ws0, header_row=1)
        wb.save(output_path)
        return output_path

    line1, line2 = run_summary
    startrow = 2
    header_row = startrow + 1
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            startrow=startrow,
        )
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    n_cols = max(len(df.columns), 1)
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=n_cols,
    )
    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=n_cols,
    )
    c1 = ws.cell(row=1, column=1, value=line1)
    c2 = ws.cell(row=2, column=1, value=line2)
    for c in (c1, c2):
        c.alignment = Alignment(wrap_text=True, vertical="top")
    _apply_excel_readability(ws, header_row=header_row)
    _apply_pubmed_hyperlinks(ws, header_row=header_row)
    wb.save(output_path)
    return output_path
